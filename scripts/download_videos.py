# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import time
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
VIDEO_DIR = ROOT / "video"
URLS_CSV = ROOT / "video_urls.csv"
OCCURRENCES_CSV = ROOT / "video_url_occurrences.csv"
MAPPING_CSV = ROOT / "video_title_url_mapping.csv"
STATUS_JSONL = ROOT / "video_download_status.jsonl"
LOG_PATH = ROOT / "download.log"
DOUYIN_SCRIPT = ROOT / "Save-DouyinVideo.ps1"

URL_RE = re.compile(r"https?://\S+")
INVALID_CHARS = set('<>:"/\\|?*')
RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}

DEFAULT_INPUT_CANDIDATES = [
    ROOT / "video_links.txt",
    ROOT / "video_links.csv",
    ROOT / "video_links.jsonl",
    ROOT / "video_links.xlsx",
    ROOT / "links.xlsx",
]
QUERY_KEYS = {"query", "question", "user_query", "用户query", "用户问题", "问题", "需求", "title", "标题"}
SEQ_KEYS = {"seq", "id", "序号", "编号"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extract video URLs from xlsx/csv/txt/jsonl inputs, download videos, "
            "and write URL/title/query mapping files."
        )
    )
    parser.add_argument(
        "--input",
        action="append",
        default=[],
        help=(
            "Input file containing video links. Supports .xlsx, .csv, .txt, .jsonl. "
            "Repeat for multiple files. If omitted, common local names are auto-detected."
        ),
    )
    parser.add_argument(
        "--query-field",
        default=None,
        help=(
            "Optional query column/key name or 1-based column index. "
            "For xlsx/csv, defaults to the second column when no header/query key is found."
        ),
    )
    parser.add_argument(
        "--url-field",
        default=None,
        help="Optional URL column/key name or 1-based column index. By default all text fields are scanned.",
    )
    return parser.parse_args()


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def log(message: str) -> None:
    line = f"[{now_iso()}] {message}"
    print(line, flush=True)
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def clean_url(url: str) -> str:
    return url.strip().rstrip("，,。;；)）]】")


def platform_of(url: str) -> str:
    low = url.lower()
    if "douyin.com" in low or "iesdouyin" in low or "douyinvod" in low:
        return "douyin"
    if "bilibili.com" in low or "b23.tv" in low:
        return "bilibili"
    return "other"


def safe_filename(name: str, max_len: int = 120) -> str:
    name = re.sub(r"\s+", " ", name or "").strip(" .\t\r\n")
    if not name:
        name = "video"
    cleaned = []
    for ch in name:
        cleaned.append("_" if ch in INVALID_CHARS or ord(ch) < 32 else ch)
    safe = "".join(cleaned).strip(" .\t\r\n")
    if not safe:
        safe = "video"
    if safe.upper() in RESERVED_NAMES:
        safe = f"{safe}_"
    if len(safe) > max_len:
        safe = safe[:max_len].strip(" .\t\r\n")
    return safe or "video"


def normalize_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def parse_column_ref(value: str | None) -> int | None:
    if not value:
        return None
    value = value.strip()
    if value.isdigit():
        index = int(value)
        return index - 1 if index > 0 else None
    return None


def cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def get_by_field(
    values: list[Any],
    header: list[str] | None,
    field: str | None,
    fallback_index: int | None = None,
    fallback_keys: set[str] | None = None,
) -> str:
    column_index = parse_column_ref(field)
    if column_index is not None and column_index < len(values):
        return cell_text(values[column_index])

    normalized_field = normalize_key(field) if field else ""
    if header:
        for idx, name in enumerate(header):
            normalized_name = normalize_key(name)
            if normalized_field and normalized_name == normalized_field and idx < len(values):
                return cell_text(values[idx])
            if fallback_keys and normalized_name in fallback_keys and idx < len(values):
                return cell_text(values[idx])

    if fallback_index is not None and fallback_index < len(values):
        return cell_text(values[fallback_index])
    return ""


def iter_urls_from_values(values: list[Any], header: list[str] | None, url_field: str | None) -> list[tuple[str, str]]:
    column_index = parse_column_ref(url_field)
    selected: list[tuple[str, Any]]
    if column_index is not None:
        selected = [(str(column_index + 1), values[column_index])] if column_index < len(values) else []
    elif url_field and header:
        normalized_field = normalize_key(url_field)
        selected = [
            (name or str(idx + 1), values[idx])
            for idx, name in enumerate(header)
            if idx < len(values) and normalize_key(name) == normalized_field
        ]
    else:
        selected = [(str(idx + 1), value) for idx, value in enumerate(values)]

    found: list[tuple[str, str]] = []
    for column, value in selected:
        text = cell_text(value)
        for match in URL_RE.finditer(text):
            found.append((column, clean_url(match.group(0))))
    return found


def looks_like_header(row: list[Any], query_field: str | None, url_field: str | None) -> bool:
    texts = [cell_text(value) for value in row]
    if any(URL_RE.search(text) for text in texts):
        return False
    normalized = {normalize_key(text) for text in texts}
    known = QUERY_KEYS | SEQ_KEYS | {"url", "link", "链接", "视频链接", "video_url", "video"}
    if normalized & known:
        return True
    named_fields = [query_field, url_field]
    return any(field and normalize_key(field) in normalized and parse_column_ref(field) is None for field in named_fields)


def extract_from_tabular(
    source_file: Path,
    rows: list[list[Any]],
    query_field: str | None,
    url_field: str | None,
) -> list[dict]:
    if not rows:
        return []
    header = [cell_text(value) for value in rows[0]] if looks_like_header(rows[0], query_field, url_field) else None
    data_rows = rows[1:] if header else rows
    row_offset = 2 if header else 1
    items: list[dict] = []

    for offset, row in enumerate(data_rows):
        row_idx = row_offset + offset
        values = list(row)
        seq = get_by_field(values, header, None, fallback_index=0, fallback_keys=SEQ_KEYS)
        query = get_by_field(values, header, query_field, fallback_index=1, fallback_keys=QUERY_KEYS)
        for column, url in iter_urls_from_values(values, header, url_field):
            items.append(
                {
                    "source_file": source_file.name,
                    "row": row_idx,
                    "column": column,
                    "seq": seq,
                    "query": query,
                    "url": url,
                    "platform": platform_of(url),
                }
            )
    return items


def extract_from_xlsx(path: Path, query_field: str | None, url_field: str | None) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    items: list[dict] = []
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        for item in extract_from_tabular(path, rows, query_field, url_field):
            item["row"] = f"{ws.title}:{item['row']}"
            items.append(item)
    return items


def extract_from_csv(path: Path, query_field: str | None, url_field: str | None) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    return extract_from_tabular(path, rows, query_field, url_field)


def strip_urls(text: str) -> str:
    return clean_text(URL_RE.sub("", text))


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip(" \t\r\n，,。;；")


def extract_from_txt(path: Path) -> list[dict]:
    items: list[dict] = []
    for row_idx, line in enumerate(path.read_text(encoding="utf-8-sig").splitlines(), start=1):
        text = line.strip()
        if not text:
            continue
        query = strip_urls(text)
        for match in URL_RE.finditer(text):
            url = clean_url(match.group(0))
            items.append(
                {
                    "source_file": path.name,
                    "row": row_idx,
                    "column": "line",
                    "seq": row_idx,
                    "query": query,
                    "url": url,
                    "platform": platform_of(url),
                }
            )
    return items


def iter_json_strings(value: Any, path: str = "$") -> list[tuple[str, str]]:
    if isinstance(value, str):
        return [(path, value)]
    if isinstance(value, dict):
        found: list[tuple[str, str]] = []
        for key, child in value.items():
            found.extend(iter_json_strings(child, f"{path}.{key}"))
        return found
    if isinstance(value, list):
        found = []
        for idx, child in enumerate(value):
            found.extend(iter_json_strings(child, f"{path}[{idx}]"))
        return found
    return []


def get_json_field(data: Any, field: str | None, fallback_keys: set[str]) -> str:
    if not isinstance(data, dict):
        return ""
    if field:
        normalized_field = normalize_key(field)
        for key, value in data.items():
            if normalize_key(key) == normalized_field and not isinstance(value, (dict, list)):
                return cell_text(value)
    for key, value in data.items():
        if normalize_key(key) in fallback_keys and not isinstance(value, (dict, list)):
            return cell_text(value)
    return ""


def extract_from_jsonl(path: Path, query_field: str | None, url_field: str | None) -> list[dict]:
    items: list[dict] = []
    for row_idx, line in enumerate(path.read_text(encoding="utf-8-sig").splitlines(), start=1):
        text = line.strip()
        if not text:
            continue
        try:
            data: Any = json.loads(text)
        except json.JSONDecodeError:
            data = {"text": text}

        seq = get_json_field(data, None, SEQ_KEYS) or row_idx
        query = get_json_field(data, query_field, QUERY_KEYS)
        selected = iter_json_strings(data)
        if url_field and isinstance(data, dict):
            normalized_field = normalize_key(url_field)
            selected = [
                (path_name, value)
                for path_name, value in selected
                if normalize_key(path_name.rsplit(".", 1)[-1]) == normalized_field
            ]
        for path_name, value in selected:
            for match in URL_RE.finditer(value):
                url = clean_url(match.group(0))
                items.append(
                    {
                        "source_file": path.name,
                        "row": row_idx,
                        "column": path_name,
                        "seq": seq,
                        "query": query or strip_urls(value),
                        "url": url,
                        "platform": platform_of(url),
                    }
                )
    return items


def default_inputs() -> list[Path]:
    return [path for path in DEFAULT_INPUT_CANDIDATES if path.exists()]


def extract_items(input_paths: list[Path], query_field: str | None, url_field: str | None) -> list[dict]:
    items: list[dict] = []
    for path in input_paths:
        if not path.exists():
            raise FileNotFoundError(f"Input file not found: {path}")
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            items.extend(extract_from_xlsx(path, query_field, url_field))
        elif suffix == ".csv":
            items.extend(extract_from_csv(path, query_field, url_field))
        elif suffix in {".txt", ".md"}:
            items.extend(extract_from_txt(path))
        elif suffix == ".jsonl":
            items.extend(extract_from_jsonl(path, query_field, url_field))
        else:
            raise ValueError(f"Unsupported input type: {path}. Use xlsx, csv, txt, md, or jsonl.")
    return items


def unique_items(items: list[dict]) -> list[dict]:
    grouped: OrderedDict[str, list[dict]] = OrderedDict()
    for item in items:
        grouped.setdefault(item["url"], []).append(item)
    rows = []
    for order, (url, occurrences) in enumerate(grouped.items(), start=1):
        rows.append(
            {
                "order": order,
                "url": url,
                "platform": platform_of(url),
                "occurrence_count": len(occurrences),
                "source_files": ";".join(str(x.get("source_file", "")) for x in occurrences),
                "source_rows": ";".join(str(x["row"]) for x in occurrences),
                "source_columns": ";".join(str(x["column"]) for x in occurrences),
                "source_seqs": ";".join(str(x["seq"]) for x in occurrences),
                "source_queries": " || ".join(x["query"] for x in occurrences),
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_url_files(items: list[dict], uniques: list[dict]) -> None:
    write_csv(
        OCCURRENCES_CSV,
        items,
        ["source_file", "row", "column", "seq", "query", "platform", "url"],
    )
    write_csv(
        URLS_CSV,
        uniques,
        [
            "order",
            "platform",
            "occurrence_count",
            "source_files",
            "source_rows",
            "source_columns",
            "source_seqs",
            "source_queries",
            "url",
        ],
    )


def completed_urls() -> set[str]:
    done = set()
    if not MAPPING_CSV.exists():
        return done
    with MAPPING_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            path = Path(row.get("file_path", ""))
            if row.get("status") == "success" and path.exists():
                done.add(row.get("url", ""))
    return done


def append_mapping(row: dict) -> None:
    fieldnames = [
        "order",
        "platform",
        "status",
        "title",
        "url",
        "file_path",
        "filename",
        "source_files",
        "source_rows",
        "source_queries",
        "started_at",
        "finished_at",
        "error",
    ]
    exists = MAPPING_CSV.exists()
    with MAPPING_CSV.open("a", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not exists:
            writer.writeheader()
        writer.writerow({k: row.get(k, "") for k in fieldnames})
    with STATUS_JSONL.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")


def run_command(args: list[str], timeout: int | None = None) -> subprocess.CompletedProcess:
    return subprocess.run(
        args,
        cwd=str(ROOT),
        text=True,
        encoding="utf-8",
        errors="replace",
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )


def get_ytdlp_title(url: str) -> str:
    proc = run_command(
        ["yt-dlp", "--dump-single-json", "--no-playlist", "--no-warnings", url],
        timeout=120,
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stdout.strip()[-2000:])
    info = json.loads(proc.stdout)
    return info.get("title") or info.get("id") or "video"


def unique_base(title: str) -> Path:
    base = VIDEO_DIR / safe_filename(title)
    if not any(VIDEO_DIR.glob(base.name + ".*")):
        return base
    for index in range(2, 1000):
        candidate = VIDEO_DIR / f"{base.name} ({index})"
        if not any(VIDEO_DIR.glob(candidate.name + ".*")):
            return candidate
    raise RuntimeError(f"Too many filename collisions for {title}")


def find_downloaded_file(base: Path) -> Path | None:
    candidates = [
        p
        for p in VIDEO_DIR.glob(base.name + ".*")
        if p.is_file() and not p.name.endswith((".part", ".ytdl", ".tmp"))
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def download_with_ytdlp(url: str) -> tuple[str, Path, str]:
    title = get_ytdlp_title(url)
    base = unique_base(title)
    output_template = str(base) + ".%(ext)s"
    args = [
        "yt-dlp",
        "--no-playlist",
        "--merge-output-format",
        "mp4",
        "--no-warnings",
        "-o",
        output_template,
        url,
    ]
    proc = run_command(args, timeout=None)
    if proc.returncode != 0:
        raise RuntimeError(proc.stdout.strip()[-3000:])
    path = find_downloaded_file(base)
    if path is None:
        raise RuntimeError("yt-dlp finished but no output file was found")
    return title, path, proc.stdout


def parse_line_value(output: str, key: str) -> str:
    prefix = key + ":"
    for line in output.splitlines():
        if line.startswith(prefix):
            return line[len(prefix) :].strip()
    return ""


def strip_douyin_id(stem: str) -> str:
    return re.sub(r"-\d{10,}$", "", stem).strip() or stem


def rename_to_title(path: Path, title: str) -> Path:
    target_base = unique_base(title)
    target = target_base.with_suffix(path.suffix.lower() or ".mp4")
    if path.resolve() == target.resolve():
        return path
    path.rename(target)
    return target


def download_with_douyin_script(url: str) -> tuple[str, Path, str]:
    if not DOUYIN_SCRIPT.exists():
        raise RuntimeError(f"Missing {DOUYIN_SCRIPT}")
    args = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(DOUYIN_SCRIPT),
        "-Url",
        url,
        "-OutputDir",
        str(VIDEO_DIR),
    ]
    proc = run_command(args, timeout=420)
    if proc.returncode != 0:
        raise RuntimeError(proc.stdout.strip()[-3000:])

    output_path = parse_line_value(proc.stdout, "Done") or parse_line_value(proc.stdout, "Output")
    safe_title = parse_line_value(proc.stdout, "SafeTitle")
    title = parse_line_value(proc.stdout, "Title")
    if not output_path:
        raise RuntimeError("Douyin script finished but did not report an output path")
    path = Path(output_path)
    if not path.exists():
        raise RuntimeError(f"Reported Douyin output does not exist: {output_path}")
    final_title = safe_title or strip_douyin_id(path.stem) or title or "douyin-video"
    final_path = rename_to_title(path, final_title)
    return title or final_title, final_path, proc.stdout


def download_one(item: dict) -> dict:
    started = now_iso()
    base_row = {
        "order": item["order"],
        "platform": item["platform"],
        "url": item["url"],
        "source_files": item.get("source_files", ""),
        "source_rows": item["source_rows"],
        "source_queries": item["source_queries"],
        "started_at": started,
        "finished_at": "",
    }
    try:
        if item["platform"] == "douyin":
            title, path, _ = download_with_douyin_script(item["url"])
        else:
            title, path, _ = download_with_ytdlp(item["url"])
        row = {
            **base_row,
            "status": "success",
            "title": title,
            "file_path": str(path),
            "filename": path.name,
            "finished_at": now_iso(),
            "error": "",
        }
        append_mapping(row)
        return row
    except Exception as exc:
        row = {
            **base_row,
            "status": "failed",
            "title": "",
            "file_path": "",
            "filename": "",
            "finished_at": now_iso(),
            "error": str(exc),
        }
        append_mapping(row)
        return row


def main() -> int:
    args = parse_args()
    input_paths = [Path(value).resolve() for value in args.input] if args.input else default_inputs()
    if not input_paths:
        raise SystemExit(
            "No input file found. Supply one or more --input files "
            "(xlsx, csv, txt, md, or jsonl), for example: python download_videos.py --input video_links.txt"
        )

    VIDEO_DIR.mkdir(exist_ok=True)
    items = extract_items(input_paths, args.query_field, args.url_field)
    uniques = unique_items(items)
    write_url_files(items, uniques)
    log("Inputs: " + ", ".join(str(path) for path in input_paths))
    log(f"Extracted {len(items)} URL occurrences, {len(uniques)} unique URLs")
    log(f"URL list: {URLS_CSV}")
    log(f"Occurrence list: {OCCURRENCES_CSV}")

    done = completed_urls()
    if done:
        log(f"Found {len(done)} already completed URLs in mapping; they will be skipped")

    total = len(uniques)
    for item in uniques:
        if item["url"] in done:
            log(f"[{item['order']}/{total}] SKIP existing {item['url']}")
            continue
        log(f"[{item['order']}/{total}] START {item['platform']} {item['url']}")
        result = download_one(item)
        if result["status"] == "success":
            log(f"[{item['order']}/{total}] DONE {result['filename']}")
            done.add(item["url"])
        else:
            log(f"[{item['order']}/{total}] FAILED {item['url']} :: {result['error'][:500]}")
        time.sleep(1)

    success = 0
    failed = 0
    if MAPPING_CSV.exists():
        with MAPPING_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                if row.get("status") == "success":
                    success += 1
                elif row.get("status") == "failed":
                    failed += 1
    log(f"Finished. success_rows={success}, failed_rows={failed}, mapping={MAPPING_CSV}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        log("Interrupted by user")
        raise
